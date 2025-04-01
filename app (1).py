# Web app logic for Coast2Coast CPR roster & upsell system

import streamlit as st
import pandas as pd
import re
import shutil
from openpyxl import load_workbook

st.set_page_config(page_title="Coast2coast X Red Cross Online Course Report Generator")

st.title("Coast2coast X Red Cross Online Course Report Generator")

# Helper functions
def extract_location(course_type):
    match = re.search(r'\(([^)]+)\)', str(course_type))
    return match.group(1).strip() if match else None

def filter_valid_courses(df):
    valid_levels = {
        "Standard First Aid & CPR/AED Level C": "Standard First Aid Blended",
        "Emergency First Aid CPR/AED Level C": "Emergency First Aid Blended",
        "CPR/AED Level C": "CPR/AED",
        "Marine Basic First Aid & CPR/AED Level C": "Marine Basic First Aid Blended"
    }
    df = df[df['Courses & Levels'].isin(valid_levels.keys())].copy()
    df['Course Level'] = df['Courses & Levels'].map(valid_levels)
    df['Course Level Code'] = df['Courses & Levels'].map({
        "Standard First Aid & CPR/AED Level C": "SFA",
        "Emergency First Aid CPR/AED Level C": "EFA",
        "CPR/AED Level C": "CPR",
        "Marine Basic First Aid & CPR/AED Level C": "MBFA"
    })
    df['Location'] = df['COURSE TYPE'].apply(extract_location)
    return df

def generate_red_cross_upload(df, course_id_df, template_path):
    output_rows = []
    unmatched = []

    shutil.copy(template_path, "Red_Cross_Upload_Filled.xlsx")
    wb = load_workbook("Red_Cross_Upload_Filled.xlsx")
    ws = wb.active

    row_num = 2
    for _, row in df.iterrows():
        match = course_id_df[
            (course_id_df['Start Date'] == row['Start']) &
            (course_id_df['Facility'].str.endswith(f"- {row['Location']}", na=False)) &
            (course_id_df['Course Level'] == row['Course Level'])
        ]
        if len(match) == 1:
            course_id = match.iloc[0]['Course ID']
            ws[f"A{row_num}"] = course_id
            ws[f"B{row_num}"] = row['First name (participant)']
            ws[f"C{row_num}"] = row['Last name (participant)']
            ws[f"D{row_num}"] = row['Email address (participant)']
            row_num += 1
        else:
            unmatched.append({
                "First Name": row['First name (participant)'],
                "Last Name": row['Last name (participant)'],
                "Email": row['Email address (participant)'],
                "Date": row['Start'],
                "Location": row['Location'],
                "Course Level": row['Course Level'],
                "Reason": "No match" if len(match) == 0 else "Multiple matches"
            })

    wb.save("Red_Cross_Upload_Filled.xlsx")
    return "Red_Cross_Upload_Filled.xlsx", pd.DataFrame(unmatched)

def generate_upsell_list(df):
    upsell_levels = {"EFA", "CPR"}
    upsell_df = df[df['Course Level Code'].isin(upsell_levels)].copy()
    return upsell_df[[
        'First name (participant)', 'Last name (participant)',
        'Email address (participant)', 'Phone (participant)',
        'Location', 'Courses & Levels'
    ]]

# File uploads
bookeo_file = st.file_uploader("Upload Bookeo Report", type=["xlsx"])
course_id_file = st.file_uploader("Upload Course ID List", type=["xlsx"])
template_file = st.file_uploader("Upload Red Cross Excel Template", type=["xlsx"])

if bookeo_file and course_id_file and template_file:
    bookeo_df = pd.read_excel(bookeo_file)
    course_df = pd.read_excel(course_id_file)

    try:
        bookeo_df_filtered = filter_valid_courses(bookeo_df)
        red_cross_path, unmatched_df = generate_red_cross_upload(bookeo_df_filtered, course_df, template_file)
        upsell_df = generate_upsell_list(bookeo_df_filtered)

        st.success("Files processed successfully!")

        with open(red_cross_path, "rb") as f:
            st.download_button("Download Red Cross Upload File", f.read(), file_name="Red_Cross_Upload_Filled.xlsx")

        st.download_button("Download Upsell Call List", upsell_df.to_csv(index=False), "upsell_call_list.csv")
        st.download_button("Download Unmatched Students Report", unmatched_df.to_csv(index=False), "unmatched_students.csv")

    except Exception as e:
        st.error(f"An error occurred: {e}")
