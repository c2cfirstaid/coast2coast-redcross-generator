
import streamlit as st
import pandas as pd
import re
import shutil
from openpyxl import load_workbook
from datetime import datetime

st.set_page_config(page_title="Coast2coast X Red Cross Online Course Report Generator")
st.title("Coast2coast X Red Cross Online Course Report Generator")

facility_mapping = {
    "Mississauga": "MI", "Ajax": "AJ", "West Ottawa": "NP", "Hamilton": "HM", "London": "LO",
    "Toronto": "TO", "North York": "NY", "Richmond Hill": "RH", "Oshawa": "OSH", "Newmarket": "NM",
    "Oakville": "OV", "Markham": "MK", "Vaughan": "VA", "Whitby": "WH", "Scarborough": "SC",
    "St Catharines": "STC", "Windsor": "WIN", "Edmonton South": "EDS", "Calgary": "CL", "Ottawa": "OT",
    "Brantford": "BF", "Belleville": "BL", "Guelph": "GL", "Burlington": "BU", "Etobicoke": "ET",
    "Kingston": "KG", "East York": "EY", "Brampton": "BR"
}

def extract_location(course_type):
    match = re.search(r'\(([^)]+)\)', str(course_type))
    return match.group(1).strip() if match else None

def filter_valid_courses(df):
    valid_levels = {
        "Standard First Aid & CPR/AED Level C": "Standard First Aid",
        "Emergency First Aid CPR/AED Level C": "Emergency First Aid",
        "CPR/AED Level C": "CPR/AED",
        "Marine Basic First Aid & CPR/AED Level C": "Marine Basic First Aid"
    }
    df = df[df['Courses & Levels'].isin(valid_levels.keys())].copy()
    df['Course Level'] = df['Courses & Levels'].map(valid_levels)
    df['Course Level Code'] = df['Courses & Levels'].map({
        "Standard First Aid & CPR/AED Level C": "SFA",
        "Emergency First Aid CPR/AED Level C": "EFA",
        "CPR/AED Level C": "CPR",
        "Marine Basic First Aid & CPR/AED Level C": "MBFA"
    })
    df['Location'] = df['COURSE TYPE'].apply(extract_location)
    return df

def generate_red_cross_upload(df, course_id_df, uploaded_template):
    output_rows = []
    unmatched = []

    # Save uploaded template to a temporary path
    temp_template_path = "temp_uploaded_template.xlsx"
    with open(temp_template_path, "wb") as f:
        f.write(uploaded_template.read())

    shutil.copy(temp_template_path, "Red_Cross_Upload_Filled.xlsx")
    wb = load_workbook("Red_Cross_Upload_Filled.xlsx")
    ws = wb.active

    df['Start'] = pd.to_datetime(df['Start']).dt.date
    course_id_df['Start Date'] = pd.to_datetime(course_id_df['Start Date']).dt.date

    row_num = 2
    for _, row in df.iterrows():
        loc_code = facility_mapping.get(row["Location"], "")
        match = course_id_df[
            (course_id_df["Start Date"] == row["Start"]) &
            (course_id_df["Facility"].str.endswith(f"- {loc_code}", na=False)) &
            (course_id_df["Course Level"].str.contains(row["Course Level"], case=False, na=False))
        ]
        if len(match) == 1:
            course_id = match.iloc[0]["Course ID"]
            ws[f"A{row_num}"] = course_id
            ws[f"B{row_num}"] = row['First name (participant)']
            ws[f"C{row_num}"] = row['Last name (participant)']
            ws[f"D{row_num}"] = row['Email address (participant)']
            row_num += 1
        else:
            unmatched.append({
                "First Name": row['First name (participant)'],
                "Last Name": row['Last name (participant)'],
                "Email": row['Email address (participant)'],
                "Date": row['Start'],
                "Location": row['Location'],
                "Course Level": row['Course Level'],
                "Reason": "No match" if len(match) == 0 else "Multiple matches"
            })

    wb.save("Red_Cross_Upload_Filled.xlsx")
    return "Red_Cross_Upload_Filled.xlsx", pd.DataFrame(unmatched)

def generate_upsell_list(df):
    upsell_levels = {"EFA", "CPR"}
    upsell_df = df[df['Course Level Code'].isin(upsell_levels)].copy()
    return upsell_df[[
        'First name (participant)', 'Last name (participant)',
        'Email address (participant)', 'Phone (participant)',
        'Location', 'Courses & Levels'
    ]]

# File uploads
bookeo_file = st.file_uploader("Upload Bookeo Report", type=["xlsx"])
course_id_file = st.file_uploader("Upload Course ID List", type=["xlsx"])
template_file = st.file_uploader("Upload Red Cross Excel Template", type=["xlsx"])

if bookeo_file and course_id_file and template_file:
    bookeo_df = pd.read_excel(bookeo_file)
    course_df = pd.read_excel(course_id_file)

    try:
        bookeo_df_filtered = filter_valid_courses(bookeo_df)
        red_cross_path, unmatched_df = generate_red_cross_upload(bookeo_df_filtered, course_df, template_file)
        upsell_df = generate_upsell_list(bookeo_df_filtered)

        st.success("Files processed successfully!")

        with open(red_cross_path, "rb") as f:
            st.download_button("Download Red Cross Upload File", f.read(), file_name="Red_Cross_Upload_Filled.xlsx")

        st.download_button("Download Upsell Call List", upsell_df.to_csv(index=False), "upsell_call_list.csv")
        st.download_button("Download Unmatched Students Report", unmatched_df.to_csv(index=False), "unmatched_students.csv")

    except Exception as e:
        st.error(f"An error occurred: {e}")
